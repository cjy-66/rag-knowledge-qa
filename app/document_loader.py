"""
文档加载与文本切分模块
支持: PDF / TXT / Markdown / Excel (.xlsx/.xls) / CSV 格式
使用 LangChain Document Loader + Text Splitter + openpyxl
"""
import csv
import os
from pathlib import Path
from typing import List

from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import (
    PyPDFLoader,
    TextLoader,
    UnstructuredMarkdownLoader,
    CSVLoader,
)

from config import config


def _load_excel(file_path: str) -> List[Document]:
    """从 Excel 文件（.xlsx / .xls）提取文本，按行分批为多个 Document"""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    filename = Path(file_path).name
    docs = []
    BATCH_ROWS = 30  # 每 30 行合并为一个 Document

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        batch_num = 0

        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
            rows_text.append(row_str)

            if len(rows_text) >= BATCH_ROWS:
                batch_num += 1
                full_text = f"[工作表: {sheet_name} / 批次 {batch_num}]\n" + "\n".join(rows_text)
                docs.append(Document(
                    page_content=full_text,
                    metadata={"source": filename, "file_path": file_path, "sheet": sheet_name},
                ))
                rows_text = []

        # 最后一批
        if rows_text:
            batch_num += 1
            full_text = f"[工作表: {sheet_name} / 批次 {batch_num}]\n" + "\n".join(rows_text)
            docs.append(Document(
                page_content=full_text,
                metadata={"source": filename, "file_path": file_path, "sheet": sheet_name},
            ))

    wb.close()
    return docs


def load_single_document(file_path: str) -> List[Document]:
    """
    根据文件类型加载单个文档

    参数:
        file_path: 文件路径

    返回:
        List[Document]: LangChain Document 对象列表

    抛出:
        ValueError: 不支持的文件类型
    """
    ext = Path(file_path).suffix.lower()

    # Excel 用 openpyxl 自定义加载
    if ext in (".xlsx", ".xls"):
        return _load_excel(file_path)

    loader_map = {
        ".pdf": PyPDFLoader,
        ".txt": TextLoader,
        ".md": UnstructuredMarkdownLoader,
        ".markdown": UnstructuredMarkdownLoader,
        ".csv": CSVLoader,
    }

    loader_cls = loader_map.get(ext)
    if loader_cls is None:
        raise ValueError(
            f"不支持的文件格式: {ext}，当前支持: .pdf, .txt, .md, .csv, .xlsx, .xls"
        )

    if ext == ".csv":
        loader = loader_cls(file_path, encoding="utf-8")
    else:
        loader = loader_cls(file_path, encoding="utf-8" if ext != ".pdf" else None)

    docs = loader.load()

    # 为每个文档块打上来源标记
    filename = Path(file_path).name
    for doc in docs:
        doc.metadata["source"] = filename
        doc.metadata["file_path"] = file_path

    return docs


def split_documents(docs: List[Document]) -> List[Document]:
    """
    对文档列表进行递归文本切分

    参数:
        docs: 原始 Document 列表

    返回:
        List[Document]: 切分后的 Document 碎片列表
    """
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=config.CHUNK_SIZE,
        chunk_overlap=config.CHUNK_OVERLAP,
        separators=["\n\n", "\n", "。", ".", "！", "？", " ", ""],
        length_function=len,
    )

    chunks = text_splitter.split_documents(docs)
    return chunks


def process_file(file_path: str) -> List[Document]:
    """
    完整处理管道：加载 → 切分

    参数:
        file_path: 文件路径

    返回:
        List[Document]: 切分好的文档块
    """
    ext = Path(file_path).suffix.lower()
    raw_docs = load_single_document(file_path)

    # Excel/CSV 已按行分批，不再进一步切分，避免产生过多碎片
    if ext in (".xlsx", ".xls", ".csv"):
        return raw_docs

    chunks = split_documents(raw_docs)
    return chunks


def get_all_uploaded_files() -> List[str]:
    """获取已上传的所有文件列表"""
    data_dir = config.DATA_DIR
    if not os.path.exists(data_dir):
        return []

    supported_exts = (".pdf", ".txt", ".md", ".markdown", ".csv", ".xlsx", ".xls")
    files = []
    for f in os.listdir(data_dir):
        if f.lower().endswith(supported_exts):
            files.append(f)
    return sorted(files)
